import pandas as pd
import openpyxl
import io
import re
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

def convert_csv_to_xlsx_safely(csv_input_data, sheet_title="Security Assessment", project_name="System Scan"):
    """
    Transforms raw CSV character matrix arrays into fully stylized Excel reports.
    Uses pure in-memory streaming to maintain 100% compatibility with non-Chromium
    WebAssembly WASM environments (like Firefox and Safari).
    
    :param csv_input_data: Can be raw string data, a bytes array, or an io stream.
    :param sheet_title: Target name for the active Excel sheet tab.
    :param project_name: Title text populated inside cell D1.
    :return: io.BytesIO buffer holding the completed spreadsheet binary.
    """
    # 1. Handle cross-browser safe input streaming boundaries
    if isinstance(csv_input_data, bytes):
        csv_str = csv_input_data.decode('utf-8', errors='ignore')
        string_stream = io.StringIO(csv_str)
    elif isinstance(csv_input_data, str):
        string_stream = io.StringIO(csv_input_data)
    else:
        string_stream = csv_input_data

    # Read dataset ensuring all network ports and host IPs are treated strictly as strings
    df = pd.read_csv(string_stream, dtype={"Host": str, "Port": str, "IP": str})
    
    # 2. Initialize in-memory OpenPyXL workbook
    output_buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    
    # Ensure gridlines are visible to the user by default
    ws.views.sheetView[0].showGridLines = True

    # 3. Create Custom Document Title Header Block
    ws.merge_cells("D1:H1")
    title_cell = ws["D1"]
    title_cell.value = f"Follow-up Plan - {project_name}"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="000000")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Leave row 2 empty for visual padding
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 15

    # 4. Generate Data Table Headers
    headers = list(df.columns)
    header_row_index = 3
    ws.row_dimensions[header_row_index].height = 26
    
    # Stylize column headers (Dark Slate Gray fill with bold white text)
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header_text in enumerate(headers, start=3):  # Start at Column C (index 3)
        cell = ws.cell(row=header_row_index, column=col_idx, value=str(header_text))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 5. Populate Data Rows
    current_row = 4
    for _, row_data in df.iterrows():
        ws.row_dimensions[current_row].height = 20
        for col_idx, value in enumerate(row_data, start=3):
            # Clean null values safely without breaking non-Chromium string string tokenizers
            cell_value = "" if pd.isna(value) else str(value)
            cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
        current_row += 1

    # 6. Apply Strict Layout Styles & Thin Black Borders
    thin_black_side = Side(style='thin', color='000000')
    grid_border = Border(left=thin_black_side, right=thin_black_side, top=thin_black_side, bottom=thin_black_side)
    
    center_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Automatically map alignments based on column content characteristics
    # Headers and numeric parameters get centered; prose remains left-aligned
    for row in ws.iter_rows(min_row=header_row_index, max_row=current_row - 1, min_col=3, max_col=len(headers) + 2):
        for cell in row:
            cell.border = grid_border
            
            # Check header string names or structural indexes to set alignments
            header_name = str(ws.cell(row=header_row_index, column=cell.column).value).lower()
            if cell.row == header_row_index:
                continue  # Skip overriding header custom alignments
                
            if any(x in header_name for x in ['port', 'protocol', 'id', 'status', 'rating', 'level', 'impact', 'likelihood', 'tier']):
                cell.alignment = center_alignment
                cell.font = Font(name="Calibri", size=10)
            else:
                cell.alignment = left_alignment
                cell.font = Font(name="Calibri", size=10)

    # 7. Dynamic Auto-Fit Width Calculator
    for col in ws.iter_cols(min_col=3, max_col=len(headers) + 2):
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.row == 1: 
                continue  # Skip merged title block to prevent extreme layout stretching
            if cell.value:
                # Handle multi-line wrapped text streams gracefully
                lines = str(cell.value).split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
        # Apply structured width constraints (min 10 characters, max 60 characters)
        ws.column_dimensions[col_letter].width = max(min(max_len + 3, 60), 10)

    # 8. Complete workbook assembly and save to buffer
    ws.auto_filter.ref = f"C3:{get_column_letter(len(headers) + 2)}{current_row - 1}"
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer

# --- Direct Execution Example Usage ---
if __name__ == "__main__":
    # Mocking client-side CSV data streams
    mock_csv_data = """Observe /Findings#,System/Asset ID,Protocol,Port,Risk Name/Observation,Risk Rating/ Level,Impact,Likelihood
V1,10.2.139.161,tcp,445,SMB Signing Not Required,Low,2,2
V2,10.2.137.153,tcp,443,SSL Certificate Expiry,Low,2,2
V3,10.2.138.175,tcp,22,SSH Server CBC Mode Ciphers,AOI,1,1"""

    print("Initializing in-memory cross-browser safe conversion processing matrix...")
    xlsx_file_stream = convert_csv_to_xlsx_safely(mock_csv_data, project_name="DWSS Verification")
    
    # Save output to disk locally
    with open("Verified_Follow_up_Plan.xlsx", "wb") as f:
        f.write(xlsx_file_stream.getbuffer())
    print("Conversion success! 'Verified_Follow_up_Plan.xlsx' generated cleanly with solid black grid borders.")
